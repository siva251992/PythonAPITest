import requests
import jsonpath
import json
import openpyxl

class Common:

    def __init__(self,Filenamepath, Sheetname):
        global wk
        global sh
        wk = openpyxl.load_workbook(Filenamepath)
        sh = wk[Sheetname]

    def fetch_row_count(self    ):
        row = sh.max_row
        return row

    def fetch_column_count(self):
        col=sh.max_column
        return col

    def fetch_key_names(self):
        c= sh.max_column
        li=[]
        for i in range(1,c+1):
            cell = sh.cell(row=1,column=i)
            #li.insert(i-1,cell.value)
            li.append(cell.value)

        return li

    def update_request_with_data(self,rowNumb,jsonrequest,keylist):
        c=sh.max_column  #4
        for i in range(1,c+1):
            cell= sh.cell(row=rowNumb,column=i)
            jsonrequest[keylist[i-1]]=cell.value #index starts from 0.

        return jsonrequest


obj = Common('C:\\Users\\esivxxx\\Desktop\\APIAutomation\\StudentData.xlsx','Sheet1')
print(obj.fetch_key_names())

