import requests
import jsonpath
import json
import openpyxl

def test_add_multiple_students():
    API_Url = 'http://thetestingworldapi.com/api/studentsDetails'
    f = open('C:\\Users\\esivxxx\\Desktop\\APIAutomation\\AddNewStudent.json', 'r')
    json_req = json.loads(f.read())
    #excelcode
    wk= openpyxl.load_workbook('C:\\Users\\esivxxx\\Desktop\\APIAutomation\\StudentData.xlsx')
    sh=wk['Sheet1']
    row = sh.max_row

    for i in range(2,row+1):
        cell_firstname = sh.cell(row=i,column=1)
        cell_midname = sh.cell(row=i, column=2)
        cell_lastname = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)
        json_req['first_name']=cell_firstname.value
        json_req['middle_name']=cell_midname.value
        json_req['last_name']=cell_lastname.value
        json_req['date_of_birth']=cell_dob.value
        response = requests.post(API_Url, json_req)
        print(response.status_code)
        print(response.text)
        assert response.status_code == 201


